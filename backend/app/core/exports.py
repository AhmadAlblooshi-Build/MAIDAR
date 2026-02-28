"""
Export services for analytics data - CSV, Excel, PDF formats.
"""

import io
import csv
import logging
from typing import List, Dict, Any
from datetime import datetime

logger = logging.getLogger(__name__)

# Optional dependencies with fallbacks
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will not be available.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed. PDF export will not be available.")


class ExportService:
    """Service for exporting data in various formats."""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "export.csv") -> bytes:
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries with export data
            filename: Name for the export file

        Returns:
            CSV file as bytes
        """
        if not data:
            return b""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        csv_bytes = output.getvalue().encode('utf-8')
        output.close()

        return csv_bytes

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        sheet_name: str = "Export",
        title: str = "MAIDAR Export"
    ) -> bytes:
        """
        Export data to Excel format with formatting.

        Args:
            data: List of dictionaries with export data
            sheet_name: Name for the Excel sheet
            title: Title for the export

        Returns:
            Excel file as bytes
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("Excel export not available. Install openpyxl.")

        if not data:
            return b""

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add title
        ws.append([title])
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="14b8a6", end_color="14b8a6", fill_type="solid")
        ws.append([])

        # Add timestamp
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"])
        ws.append([])

        # Add headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Format header row
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0d9488", end_color="0d9488", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_data in data:
            ws.append(list(row_data.values()))

        # Auto-adjust column widths
        for column in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(column)

            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()

        return excel_bytes

    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        title: str = "MAIDAR Export",
        subtitle: str = None
    ) -> bytes:
        """
        Export data to PDF format with professional formatting.

        Args:
            data: List of dictionaries with export data
            title: Title for the PDF
            subtitle: Optional subtitle

        Returns:
            PDF file as bytes
        """
        if not PDF_AVAILABLE:
            raise RuntimeError("PDF export not available. Install reportlab.")

        if not data:
            return b""

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#14b8a6'),
            spaceAfter=12,
            alignment=1  # Center alignment
        )
        elements.append(Paragraph(title, title_style))

        # Subtitle
        if subtitle:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.grey,
                spaceAfter=20,
                alignment=1
            )
            elements.append(Paragraph(subtitle, subtitle_style))

        elements.append(Spacer(1, 0.2*inch))

        # Metadata
        meta_data = [
            ["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["Records:", str(len(data))],
        ]
        meta_table = Table(meta_data, colWidths=[1.5*inch, 4*inch])
        meta_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.3*inch))

        # Data table
        headers = list(data[0].keys())
        table_data = [headers]

        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        # Calculate column widths based on content
        available_width = 7*inch
        col_widths = [available_width / len(headers)] * len(headers)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#0d9488')),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))

        elements.append(table)

        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        elements.append(Paragraph("MAIDAR - Human Risk Intelligence Platform", footer_style))
        elements.append(Paragraph("© 2026 MAIDAR. All rights reserved.", footer_style))

        # Build PDF
        doc.build(elements)
        pdf_bytes = output.getvalue()
        output.close()

        return pdf_bytes


# Singleton instance
export_service = ExportService()


# Export type helpers
def export_employees(employees: List[Dict], format: str = "csv") -> bytes:
    """Export employee data."""
    if format == "csv":
        return export_service.export_to_csv(employees, "employees.csv")
    elif format == "excel":
        return export_service.export_to_excel(employees, "Employees", "Employee Export")
    elif format == "pdf":
        return export_service.export_to_pdf(employees, "Employee Export", f"Total Employees: {len(employees)}")
    else:
        raise ValueError(f"Unsupported format: {format}")


def export_risk_scores(risk_data: List[Dict], format: str = "csv") -> bytes:
    """Export risk score data."""
    if format == "csv":
        return export_service.export_to_csv(risk_data, "risk_scores.csv")
    elif format == "excel":
        return export_service.export_to_excel(risk_data, "Risk Scores", "Risk Score Export")
    elif format == "pdf":
        return export_service.export_to_pdf(risk_data, "Risk Score Export", f"Total Records: {len(risk_data)}")
    else:
        raise ValueError(f"Unsupported format: {format}")


def export_simulation_results(results: List[Dict], format: str = "csv") -> bytes:
    """Export simulation results."""
    if format == "csv":
        return export_service.export_to_csv(results, "simulation_results.csv")
    elif format == "excel":
        return export_service.export_to_excel(results, "Results", "Simulation Results Export")
    elif format == "pdf":
        return export_service.export_to_pdf(results, "Simulation Results", f"Total Results: {len(results)}")
    else:
        raise ValueError(f"Unsupported format: {format}")
